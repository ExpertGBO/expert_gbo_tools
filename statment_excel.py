from itertools import product
import psycopg2
from psycopg2 import sql
import json
from datetime import date, datetime, time
import openpyxl
import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
from openpyxl.worksheet.page import PageMargins
from psycopg2.extras import RealDictCursor

# Параметры подключения к базе данных
db_params = {
    'host': '192.168.2.111',
    'database': 'expert_pool',
    'user': 'admin',
    'password': 'Eon7a27i1ZczZ59onvFQ'
}

# Даты для фильтрации
start_date = '2024-11-25'
end_date = '2024-12-01'
ppto = ""

# Путь к шаблону и директории для сохранения файлов
template_path = os.path.expanduser('Акт сверки Шаблон.xlsx')
output_dir = os.path.expanduser('~/Desktop/счета')

# Создание директории для сохранения файлов, если она не существует
if not os.path.exists(output_dir):
    os.makedirs(output_dir)


# Кастомный JSON-энкодер для обработки типов date, datetime и time
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        elif isinstance(obj, time):
            return obj.strftime('%H:%M:%S')
        return super().default(obj)

def get_single_value(value):
    if isinstance(value, tuple):
        return value[0]
    return value

# Функция для выполнения запроса и получения данных из базы данных
def fetch_data(db_params, start_date, end_date, ppto):
    try:
        # Подключение к базе данных
        connection = psycopg2.connect(**db_params)
        cursor = connection.cursor()

        # SQL-запрос с условием для ppto, если она не пуста
        if ppto:
            query = sql.SQL("""
                          SELECT *
                          FROM account_cp_view
                          WHERE doc_date BETWEEN %s AND %s AND ppto = %s AND doc_status = 'Подписан' ORDER BY doc_date ASC
                      """)
            cursor.execute(query, (start_date, end_date, ppto))
        else:
            query = sql.SQL("""
                          SELECT *
                          FROM account_cp_view
                          WHERE doc_date BETWEEN %s AND %s AND doc_status = 'Подписан' ORDER BY doc_date ASC
                      """)
            cursor.execute(query, (start_date, end_date))

        # Извлечение данных
        rows = cursor.fetchall()

        # Получение имен столбцов
        colnames = [desc[0] for desc in cursor.description]

        # Создание вложенного словаря для сортировки данных
        result_dict = {}
        unique_doc_numbers = set()  # Множество для отслеживания уникальных doc_number

        for row in rows:
            row_dict = dict(zip(colnames, row))
            doc_number = row_dict['doc_number']
            ppto_key = row_dict['recipient']

            if ppto_key not in result_dict:
                result_dict[ppto_key] = {'conclusions': [], 'protocols': []}

            if doc_number in unique_doc_numbers:
                # Документ уже добавлен, пропускаем
                continue

            if doc_number.startswith("ПТЭ"):
                result_dict[ppto_key]['conclusions'].append(row_dict)
            elif doc_number.startswith("ПБ"):
                result_dict[ppto_key]['protocols'].append(row_dict)



            # Добавляем doc_number в множество
            unique_doc_numbers.add(doc_number)

        # Дополнительный запрос для данных с corr_price IS NOT NULL
        if ppto:
            corr_query = sql.SQL("""
                              SELECT *
                              FROM account_cp_view
                              WHERE corr_price IS NOT NULL 
                                AND corr_date BETWEEN %s AND %s
                                AND ppto = %s ORDER BY doc_date ASC
                          """)
            cursor.execute(corr_query, (start_date, end_date, ppto))
        else:
            corr_query = sql.SQL("""
                              SELECT *
                              FROM account_cp_view
                              WHERE corr_price IS NOT NULL 
                                AND corr_date BETWEEN %s AND %s ORDER BY doc_date ASC
                          """)
            cursor.execute(corr_query, (start_date, end_date))

        corr_rows = cursor.fetchall()

        # Обновление result_dict данными из дополнительного запроса
        for row in corr_rows:
            row_dict = dict(zip(colnames, row))
            doc_number = row_dict['doc_number']
            ppto_key = row_dict['recipient']

            if ppto_key not in result_dict:
                result_dict[ppto_key] = {'conclusions': [], 'protocols': []}

            if doc_number in unique_doc_numbers:
                # Документ уже добавлен, пропускаем
                continue

            if doc_number.startswith("ПТЭ"):
                result_dict[ppto_key]['conclusions'].append(row_dict)
            elif doc_number.startswith("ПБ"):
                result_dict[ppto_key]['protocols'].append(row_dict)
            # Добавляем doc_number в множество
            unique_doc_numbers.add(doc_number)

        # Закрытие курсора и соединения
        cursor.close()
        connection.close()

        return result_dict
    except Exception as error:
        print(f"Ошибка при выполнении запроса: {error}")
        if connection:
            cursor.close()
            connection.close()

# Функция для получения JSON данных из таблицы
def fetch_json_data(db_params):
    try:
        # Подключение к базе данных
        connection = psycopg2.connect(**db_params)
        cursor = connection.cursor()

        # SQL-запрос для получения JSON данных
        query = sql.SQL("""
            SELECT json
            FROM json_price_list WHERE id = 72 
        """)

        # Выполнение запроса
        cursor.execute(query)

        # Извлечение данных
        result = cursor.fetchone()

        if result:
            json_data = result[0]
            # Если json_data уже является словарем, просто передаем его
            if isinstance(json_data, str):
                data_dict = json.loads(json_data)  # Преобразуем JSON строку в словарь
            else:
                data_dict = json_data  # Уже словарь
        else:
            data_dict = {}  # Возвращаем пустой словарь, если данных нет

        # Закрытие курсора и соединения
        cursor.close()
        connection.close()

        return data_dict
    except Exception as error:
        print(f"Ошибка при выполнении запроса: {error}")
        if connection:
            cursor.close()
            connection.close()
        return {}

def new_price (db_params, ppto):

        # Подключение к базе данных
        connection = psycopg2.connect(**db_params)
        cursor = connection.cursor(cursor_factory=RealDictCursor)

        # SQL-запрос для получения JSON данных
        query = f"SELECT * from ppto WHERE name = '{ppto}'"

        # Выполнение запроса
        cursor.execute(query)

        # Извлечение данных
        result = cursor.fetchone()
        if result:
            n_price = result['new_price']
            sale = result['sale']
            cursor.close()
            connection.close()
            return n_price,  sale
        else:
            return False, False




# Вызов функции и получение результата
data = fetch_data(db_params, start_date, end_date, ppto)
json_data_dict = fetch_json_data(db_params)


# Сортировка словаря по ключу ppto
sorted_data = dict(sorted(data.items()))


# Функция для очистки имени файла от недопустимых символов
def clean_filename(filename):
    # Удаление недопустимых символов
    cleaned_filename = re.sub(r'[<>:"/\\|?*]', '', filename)
    return cleaned_filename

# Функция для форматирования даты в нужный формат
def format_date(date_obj):
    """Форматирует дату в строку формата 'ДД.ММ.ГГГГ'."""
    if isinstance(date_obj, str):
        try:
            # Пробуем распарсить строку в формате 'YYYY-MM-DD'
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
        except ValueError:
            # Если не удалось, пробуем другой формат или сообщаем об ошибке
            raise ValueError(f"Неподдерживаемый формат даты: {date_obj}")

    elif isinstance(date_obj, datetime):
        # Если объект datetime, преобразуем в формат 'DD.MM.YYYY'
        return date_obj.strftime('%d.%m.%Y')

    elif isinstance(date_obj, date):
        # Если объект date, преобразуем в формат 'DD.MM.YYYY'
        return date_obj.strftime('%d.%m.%Y')

    else:
        raise TypeError(f"Ожидалась строка или объект даты, но получен {type(date_obj)}")

    # Если date_obj был строкой и успешно преобразован, возвращаем форматированную строку
    return date_obj.strftime('%d.%m.%Y')

def set_row_height(ws, row, min_height=15):
    # Пробуем определить высоту строки по содержимому
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[row]) -5
    # Примерное правило для расчета высоты строки
    height = max(min_height, max_length )  # Параметр 10 можно настроить в зависимости от шрифта и размера
    if ws.row_dimensions[row].height < height:
        ws.row_dimensions[row].height = height


# Функция для создания Excel-файла на основе шаблона
def create_excel_file(template_path, output_path, ppto, data, n_price, sale):

    # Загрузка шаблона
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    price_sheet = workbook["Прайс"]
    price_sheet_1 = workbook["Ведомость"]

    # Получение объединенных ячеек
    merged_cells = sheet.merged_cells.ranges

    def write_value(cell, value):
        """Записывает значение в верхнюю левую ячейку объединенной области."""
        for merged_range in merged_cells:
            if cell.coordinate in merged_range:
                top_left_cell = price_sheet_1.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = value
                return

        cell.value = value

    # Создание объекта рамки
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Переменные для накопления общей стоимости
    total_conclusions_price = 0
    total_protocols_price = 0

    # Настройка шрифта и выравнивания
    custom_font = Font(name='Times New Roman', size=14, bold=True, italic=False, color='000000')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Форматирование дат
    formatted_start_date = format_date(start_date)
    formatted_end_date = format_date(end_date)

    if n_price:
        # Шапка прайс
        n_price_2 = "(Новый прайс)"
        price_sheet.cell(row=1,column=1).value = "Прайс-лист\nна проведение работ по оценке соответствия ТС" + " (" + ppto + ")"
        cell = price_sheet.cell(row=1, column=1)
        cell.font = custom_font  # Применение шрифта
        # Пройтись по ячейкам в столбце B и заполнить данные
        start_row = 5
        current_row = start_row
        price_sp = False

        for i in range(6, 120):
            cell_value = price_sheet.cell(row=i, column=7).value

            # Проверка перед первым вызовом upper()
            if cell_value is not None and isinstance(cell_value, str) and cell_value != "":
                if cell_value.upper() in json_data_dict["Общий"]:
                    price_sheet.cell(row=i, column=3).value = str(json_data_dict["Общий"][cell_value.upper()]['ПЗ Легковые'])
                    price_sheet.cell(row=i, column=4).value = str(json_data_dict["Общий"][cell_value.upper()]['ПБ Легковые'])
                    price_sheet.cell(row=i, column=5).value = str(json_data_dict["Общий"][cell_value.upper()]['ПЗ Грузовые'])
                    price_sheet.cell(row=i, column=6).value = str(json_data_dict["Общий"][cell_value.upper()]['ПБ Грузовые'])

            try:
                # Проверка перед вызовом upper() внутри try
                if cell_value is not None and isinstance(cell_value, str) and cell_value != "":
                    if cell_value.upper() in json_data_dict[ppto]:
                        if json_data_dict[ppto][cell_value.upper()]['ПЗ Легковые'] is not None:
                            price_sheet.cell(row=i, column=3).value = str(
                                json_data_dict[ppto][cell_value.upper()]['ПЗ Легковые'])
                            cell = price_sheet.cell(row=i, column=3)
                            cell.alignment = center_alignment
                        if json_data_dict[ppto][cell_value.upper()]['ПБ Легковые'] is not None:
                            price_sheet.cell(row=i, column=4).value = str(
                                json_data_dict[ppto][cell_value.upper()]['ПБ Легковые'])
                            cell = price_sheet.cell(row=i, column=4)
                            cell.alignment = center_alignment
                        if json_data_dict[ppto][cell_value.upper()]['ПЗ Грузовые'] is not None:
                            price_sheet.cell(row=i, column=5).value = str(
                                json_data_dict[ppto][cell_value.upper()]['ПЗ Грузовые'])
                            cell = price_sheet.cell(row=i, column=5)
                            cell.alignment = center_alignment
                        if json_data_dict[ppto][cell_value.upper()]['ПБ Грузовые'] is not None:
                            price_sheet.cell(row=i, column=6).value = str(
                                json_data_dict[ppto][cell_value.upper()]['ПБ Грузовые'])
                            cell = price_sheet.cell(row=i, column=6)
                            cell.alignment = center_alignment
            except KeyError:
                pass  # Игнорируем KeyError, если ключ не найден
    else:
        workbook.remove(price_sheet)
        n_price_2 = ""

    if sale:
        n_price_2 = n_price_2 + " + скидка " + sale
        price_sheet.cell(row=1,
                         column=1).value = "Прайс-лист\nна проведение работ по оценке соответствия ТС" + " (" + ppto + ")" + " + скидка " + sale



    custom_font = Font(name='Times New Roman', size=20, bold=False, italic=False, color='000000')

    # Запись текста в ячейку A2
    text = f"заключений и протоколов, выданных в период с {formatted_start_date} по {formatted_end_date} {ppto}"
    cell_a2 = price_sheet_1.cell(row=2, column=1)
    write_value(cell_a2, text)
    cell_a2.font = custom_font
    cell_a2.alignment = center_alignment

    custom_font = Font(name='Times New Roman', size=14, bold=False, italic=False, color='000000')


    # Начальные строки для заключений и протоколов
    row_num_z = 6
    row_num_p = 6

    # Определяем цвет заливки
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    categories = {"N2", "N2G", "N3", "N3G", "M2", "M2G", "M3", "M4G", "O3", "O4"}



    # Заполнение данных заключений
    for entry in data['conclusions']:
        if entry['corr_price'] is not None:
            price = entry['corr_price']
            doc_date = str(format_date(entry['corr_date'])) + "\nИсправление"
            changes = entry['new_change'] + "\n" + entry['corr_price_comment']


        else:
            price = entry['price']
            doc_date = str(format_date(entry['doc_date']))
            changes = entry['change']

# Колличество переоборудования
        change_count = changes.count('+')

        if changes != 'Дем. ГБО+ГБО' and changes != 'ГБО+Дем. ГБО' and changes != 'ГБО+ТСУ' and changes != 'ТСУ+ГБО' and entry['corr_price'] is None:
            if n_price == True and 1 <= change_count <= 3:
                changes = changes + ' (два и более)'
            elif n_price == True and 4 <= change_count <= 8:
                changes = changes + ' (пять и более)'
            elif n_price == True and change_count >= 9:
                changes = changes + ' (десять и более)'


        cells_to_fill = {
            7: entry['doc_number'],
            6: entry['ppto'],
            2: doc_date,
            8: entry['grz'],
            3: changes,
            9: entry['fio'],
            5: entry['category'],
            4: price,
            1: row_num_z - 5
        }



        category = entry['category']


        # Определяем, нужно ли выделять строку
        fill_color = green_fill if category in categories else None

        if fill_color:
            cell = price_sheet_1.cell(row=row_num_z, column=5)
            cell.fill = fill_color

        price = ""
        changes = ""
        doc_date = ""

        for col, value in cells_to_fill.items():
            cell = price_sheet_1.cell(row=row_num_z, column=col)
            write_value(cell, value)
            cell.font = custom_font
            cell.alignment = center_alignment
            cell.border = border



        price_sheet_1.row_dimensions[row_num_z].height = 40

        # Пример вызова
        set_row_height(price_sheet_1, row_num_z)

        row_num_z += 1

        if entry['corr_price'] is not None:
            price = entry.get('corr_price')  # Заменяет None на 0, если price не существует

        else:
            price = entry.get('price')



        total_conclusions_price = total_conclusions_price + price

    custom_font = Font(name='Times New Roman', size=20, bold=True, italic=False, color='000000')

    # Итоговая строка для заключений
    cell = price_sheet_1.cell(row=row_num_z, column=3)
    write_value(cell, "Итого:")
    cell.font = custom_font
    cell.alignment = center_alignment
    cell.border = border
    price_sheet_1.row_dimensions[row_num_z].height = 40

    cell = price_sheet_1.cell(row=row_num_z, column=4)
    write_value(cell, total_conclusions_price)
    cell.font = custom_font
    cell.alignment = center_alignment
    cell.border = border

    custom_font = Font(name='Times New Roman', size=14, bold=False, italic=False, color='000000')

    # Заполнение данных протоколов
    for entry in data['protocols']:

        if entry['corr_price'] is not None:
            price = entry['corr_price']  # Удалён лишний запятая, чтобы не создавать кортеж
            doc_date = str(format_date(entry['corr_date'])) + "\nИсправление"
            changes = entry['change'] + "\n" + entry['corr_price_comment']
            if entry['alien_lab'] == True and n_price ==True:
                changes = changes + " (Заключение сторонней ИЛ доплата 500р)"
        else:
            price = entry['price']
            doc_date = str(format_date(entry['doc_date']))
            changes = entry['change']
            if entry['alien_lab'] == True and n_price ==True:
                changes = changes + " (Заключение сторонней ИЛ доплата 500р)"

        # Колличество переоборудования
        change_count = changes.count('+')



        # Добавить шифр по колличеству переоборудований
        if changes != 'Дем. ГБО+ГБО' and changes != 'ГБО+Дем. ГБО' and changes != 'ГБО+ТСУ' and changes != 'ТСУ+ГБО' and entry['corr_price'] is None:
            if n_price == True and 1 <= change_count <= 3:
                changes = changes + ' (два и более)'
            elif n_price == True and 4 <= change_count <= 8:
                changes = changes + ' (пять и более)'
            elif n_price == True and change_count >= 9:
                changes = changes + ' (десять и более)'


        cells_to_fill = {
            16: entry['doc_number'],
            15: entry['ppto'],
            11: doc_date,
            17: entry['grz'],
            12: changes,
            18: entry['fio'],
            14: entry['category'],
            13: price,
            10: row_num_p - 5
        }

        category = entry['category']



        # Определяем, нужно ли выделять строку
        fill_color = green_fill if category in categories else None

        if entry['corr_price'] is not None:
            price = entry.get('corr_price')  # Заменяет None на 0, если price не существует

        else:
            price = entry.get('price')



        total_protocols_price = total_protocols_price + price

        if fill_color:
            cell = price_sheet_1.cell(row=row_num_p, column=14)
            cell.fill = fill_color


        for col, value in cells_to_fill.items():
            cell = price_sheet_1.cell(row=row_num_p, column=col)
            write_value(cell, value)
            cell.font = custom_font
            cell.alignment = center_alignment
            cell.border = border

        price_sheet_1.row_dimensions[row_num_p].height = 40

        # Пример вызова
        set_row_height(price_sheet_1, row_num_p)


        row_num_p += 1

    custom_font = Font(name='Times New Roman', size=20, bold=True, italic=False, color='000000')

    # Итоговая строка для протоколов
    cell = price_sheet_1.cell(row=row_num_p, column=12)
    write_value(cell, "Итого:")
    cell.font = custom_font
    cell.alignment = center_alignment
    cell.border = border

    cell = price_sheet_1.cell(row=row_num_p, column=13)
    write_value(cell, total_protocols_price)
    cell.font = custom_font
    cell.alignment = center_alignment
    cell.border = border

    price_sheet_1.row_dimensions[row_num_p].height = 40

    # Определение последней строки для итоговой строки стоимости
    row_last = max(row_num_z, row_num_p)

    # Итоговая строка стоимости заключений и протоколов
    price_sheet_1.row_dimensions[row_last + 2].height = 40
    range_to_merge = f'J{row_last+2}:Q{row_last+2}'
    (price_sheet_1.merge_cells(range_to_merge))

    custom_font_price = Font(name='Times New Roman', size=20, bold=True, italic=False, color='000000')
    cell = price_sheet_1.cell(row=row_last+2, column=10)
    write_value(cell, "Итого стоимость заключений и протоколов:")
    cell.font = custom_font_price
    cell.alignment = center_alignment
    cell.border = border
    for col in range(11, 18):
        cell = price_sheet_1.cell(row=row_last + 2, column=col)
        cell.border = border
    cell = price_sheet_1.cell(row=row_last + 2, column=18)
    write_value(cell, total_protocols_price+total_conclusions_price)
    cell.font = custom_font_price
    cell.alignment = center_alignment
    cell.border = border



    # Подписи
    price_sheet_1.row_dimensions[row_last + 5].height = 40
    price_sheet_1.row_dimensions[row_last + 6].height = 40
    range_to_merge = f'A{row_last + 5}:I{row_last + 5}'
    price_sheet_1.merge_cells(range_to_merge)
    range_to_merge = f'A{row_last + 6}:I{row_last + 6}'
    price_sheet_1.merge_cells(range_to_merge)
    range_to_merge = f'J{row_last + 5}:R{row_last + 5}'
    price_sheet_1.merge_cells(range_to_merge)
    range_to_merge = f'J{row_last + 6}:R{row_last + 6}'
    price_sheet_1.merge_cells(range_to_merge)

    custom_font_il = Font(name='Times New Roman', size=20, bold=False, italic=False, color='000000')
    cell = price_sheet_1.cell(row=row_last + 5, column=1)
    write_value(cell, "ООО Эксперт ГБО, Руководитель ИЛ  Курочкин А.И.")
    cell.font = custom_font_il




    cell = price_sheet_1.cell(row=row_last + 6, column=1)
    write_value(cell, "М.П.")
    cell.font = custom_font_il


    cell = price_sheet_1.cell(row=row_last + 5, column=10)
    write_value(cell, ppto)
    cell.font = custom_font_il


    cell = price_sheet_1.cell(row=row_last + 6, column=10)
    write_value(cell, "М.П.")
    cell.font = custom_font_il

    # Устанавливаем размер бумаги и ориентацию
    price_sheet_1.page_setup.paperSize = price_sheet_1.PAPERSIZE_A4  # Размер бумаги A4
    price_sheet_1.page_setup.orientation = price_sheet_1.ORIENTATION_LANDSCAPE  # Альбомная ориентация

    # Настройка масштабирования
    price_sheet_1.page_setup.fitToWidth = 1  # Масштабировать до 1 страницы по ширине
    price_sheet_1.page_setup.fitToHeight = 1  # Не масштабировать по высоте (можно изменить на 1, если нужно, чтобы все поместилось на одну страницу)

    price_sheet_1.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)



    # Форматирование имени файла
    formatted_start_date = format_date(start_date)
    formatted_end_date = format_date(end_date)
    cleaned_ppto = clean_filename(ppto)
    filename = f"Акт сверки Эксперт ГБО {cleaned_ppto} c {formatted_start_date} по {formatted_end_date} {n_price_2}.xlsx"
    output_path = os.path.join(output_dir, filename)

    # Сохранение в новый файл
    workbook.save(output_path)


# Создание Excel-файлов для каждого ppto
for ppto, ppto_data in sorted_data.items():
    n_price, sale = new_price(db_params, ppto)
    if ppto == "ООО Феникс (Брокер)":
        n_price = True

    print(ppto)
    print(n_price)


    create_excel_file(template_path, None, ppto, ppto_data, n_price, sale)


print("Файлы созданы и сохранены в папку счета на рабочем столе.")